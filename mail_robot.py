# encoding=utf-8
import datetime
import smtplib
import time
import traceback
from email.header import Header
from email.mime.application import MIMEApplication  # 用于添加附件
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import parseaddr, formataddr

from openpyxl import load_workbook
from pandas import ExcelFile
from pymssql import connect

from ding_robot import ding


def get_datas():
    try:
        # 从数据库中读取目标数据
        cities = ['RAPDB_HN_HaiKou_PRC', 'RAPDB_SD_HeZe_PRC', 'RAPDB_SC_MianYang_PRC']
        datas = []
        # 跟数据库建立连接
        conn = connect(host='171.217.92.221:12223',
                       user='CAP_HeWanTing',
                       password='9JTToK*pExMgh7e$@U#UdcoMiApq1oLI',
                       database='RAPDB_HN_HaiKou_TM',
                       autocommit=True)
        # 使用 cursor() 方法创建一个游标对象 cursor
        cur = conn.cursor()
        # 使用 execute() 方法执行 SQL
        # 获取所需要的数据
        for item in cities:
            cur.execute("SELECT COUNT(1) FROM {}..t_p_batchinfo".format(item))
            count1 = int(cur.fetchone()[0])
            cur.execute("SELECT COUNT(1) FROM {}..t_p_building".format(item))
            count2 = int(cur.fetchone()[0])
            cur.execute("SELECT COUNT(1) FROM {}..t_p_rooms".format(item))
            count3 = int(cur.fetchone()[0])
            datas.append((count1, count2, count3))
        # 关闭连接
        cur.close()
        # 返回所需的数据
        return datas
    except Exception as e:
        print('连接数据库失败！')
        print(traceback.format_exc())
        ding('连接数据库失败！' + str(e))


def get_excel(data, file):
    try:
        # 打开已有表格
        wb = load_workbook('template.xlsx')
        ws = wb['Sheet1']
        ws['C4'] = data[0][0]
        ws['D4'] = data[0][1]
        ws['E4'] = data[0][2]

        ws['C5'] = data[1][0]
        ws['D5'] = data[1][1]
        ws['E5'] = data[1][2]

        ws['C6'] = data[2][0]
        ws['D6'] = data[2][1]
        ws['E6'] = data[2][2]

        date = time.strftime("%Y/%m/%d", time.localtime())
        ws['B10'] = date + '-' + date
        ws['B11'] = date + '-' + date
        ws['B12'] = date + '-' + date

        wb.save("./tables/{}".format(file))
        # 返回生成的excel
        # return wb
    except Exception as e:
        print('写入数据失败！')
        print(traceback.format_exc())
        ding('写入数据失败！' + str(e))


def get_content(file_path):
    try:
        xd = ExcelFile(file_path, engine='openpyxl')
        df = xd.parse()
        print('df:', df)
        content = df.to_html(header=False, index=False, na_rep='').replace('<td>采集情况</td>',
                                                                           '<td colspan="6" bgcolor=yellow class="title">采集情况</td>').replace(
            '<td>处理情况</td>', '<td colspan="9" bgcolor=yellow class="title">处理情况</td>').replace('<td></td>', '')

        head = \
            """
            <head>
                <meta charset="utf-8">
                <STYLE TYPE="text/css" MEDIA=screen>
                    .title {
                        font-weight: bold;
                    }
                    table.dataframe {
                        border-collapse: collapse;
                        border: 2px solid #a19da2;
                        /*居中显示整个表格*/
                        margin: auto;
                    }
                    table.dataframe thead {
                        border: 2px solid #91c6e1;
                        background: #f1f1f1;
                        padding: 10px 10px 10px 10px;
                        color: #333333;
                    }
                    table.dataframe tbody {
                        border: 2px solid #91c6e1;
                        padding: 10px 10px 10px 10px;
                    }
                    table.dataframe tr {
                    }
                    table.dataframe th {
                        vertical-align: top;
                        font-size: 14px;
                        padding: 10px 10px 10px 10px;
                        color: #105de3;
                        font-family: arial;
                        text-align: center;
                    }
                    table.dataframe td {
                        text-align: center;
                        padding: 10px 10px 10px 10px;
                    }
                    body {
                        font-family: 宋体;
                    }
                    h1 {
                        color: #5db446
                    }
                    div.header h2 {
                        color: #0002e3;
                        font-family: 黑体;
                    }
                    div.content h2 {
                        text-align: center;
                        font-size: 28px;
                        text-shadow: 2px 2px 1px #de4040;
                        color: #fff;
                        font-weight: bold;
                        background-color: #008eb7;
                        line-height: 1.5;
                        margin: 20px 0;
                        box-shadow: 10px 10px 5px #888888;
                        border-radius: 5px;
                    }
                    h3 {
                        font-size: 22px;
                        background-color: rgba(0, 2, 227, 0.71);
                        text-shadow: 2px 2px 1px #de4040;
                        color: rgba(239, 241, 234, 0.99);
                        line-height: 1.5;
                    }
                    h4 {
                        color: #e10092;
                        font-family: 楷体;
                        font-size: 20px;
                        text-align: center;
                    }
                    td img {
                        /*width: 60px;*/
                        max-width: 300px;
                        max-height: 300px;
                    }
                </STYLE>
            </head>
            """
        body = \
            """
            <body>
            <div class="content">
                <!--正文内容-->
                <h2> </h2>
                <div>
                    <h4></h4>
                    {df_html}
                </div>
                <p style="text-align: center">
                </p>
            </div>
            </body>
            """.format(df_html=content)
        html_msg = "<html>" + head + body + "</html>"
        html_msg = html_msg.replace('\n', '').encode("utf-8")
        return html_msg
    except Exception as e:
        print('转换为html正文失败！')
        print(traceback.format_exc())
        ding('转换为html正文失败！' + str(e))


def create_email(email_from, email_to, email_subject, email_text, annex_path, annex_name):
    try:
        # 输入发件人昵称、收件人昵称、主题，正文，附件地址,附件名称生成一封邮件
        # 生成一个空的带附件的邮件实例
        message = MIMEMultipart()
        # 将正文以text的形式插入邮件中
        message.attach(MIMEText(email_text, 'html', 'utf-8'))
        # 生成发件人名称（这个跟发送的邮件没有关系）
        # message['From'] = Header(email_from, 'utf-8')
        message['From'] = formataddr(parseaddr(email_from))
        # 生成收件人名称（这个跟接收的邮件也没有关系）
        # message['To'] = Header(email_to, 'utf-8')
        message['To'] = formataddr(parseaddr(email_to))
        # 生成邮件主题
        message['Subject'] = Header(email_subject, 'utf-8')
        # 读取附件的内容
        att = MIMEApplication(open(annex_path, 'rb').read(), 'base64')
        att["Content-Type"] = 'application/octet-stream'
        # 生成附件的名称
        att.add_header('content-disposition', 'attachment',
                       filename=('utf-8', '', annex_name))  # 注意：此处basename要转换为gbk编码，否则中文会有乱码。
        # 将附件内容插入邮件中
        message.attach(att)
        print('邮件创建成功！')
        # 返回邮件
        return message
    except Exception as e:
        print('邮件创建失败！')
        print(traceback.format_exc())
        ding('邮件创建失败！' + str(e))


def send_email(sender, password, receiver, msg):
    # 一个输入邮箱、密码、收件人、邮件内容发送邮件的函数
    try:
        # 找到你的发送邮箱的服务器地址，已加密的形式发送
        server = smtplib.SMTP_SSL("smtp.qq.com")  # 发件人邮箱中的SMTP服务器
        server.ehlo()
        # 登录你的账号
        server.login(sender, password)  # 括号中对应的是发件人邮箱账号、邮箱密码
        # 发送邮件
        server.sendmail(sender, receiver, msg.as_string())  # 括号中对应的是发件人邮箱账号、收件人邮箱账号（是一个列表）、邮件内容
        print("邮件发送成功")
        server.quit()  # 关闭连接
    except Exception as e:
        print("邮件发送失败")
        print(traceback.format_exc())
        ding('邮件发送失败！' + str(e))


def main():
    print(datetime.datetime.now())
    date = time.strftime("%Y%m%d", time.localtime())

    # 生成数据
    my_data = get_datas()

    # 文件名称
    my_file_name = '城市采集表数据检查_' + date + '.xlsx'
    # my_file_name = 'test.xlsx'
    print('my_file_name:', my_file_name)

    # 生成excel
    get_excel(my_data, my_file_name)

    # 文件路径
    new_file_path = './tables/' + my_file_name

    my_email_from = '1659026978@qq.com'
    my_email_to = 'reallyinfo'
    # 邮件标题
    my_email_subject = '城市采集表数据检查_' + date + '-贺琬婷'
    # 邮件正文
    my_email_text = get_content(new_file_path)
    # 附件地址
    my_annex_path = new_file_path
    # 附件名称
    my_annex_name = my_file_name
    # 生成邮件
    my_msg = create_email(my_email_from, my_email_to, my_email_subject,
                          my_email_text, my_annex_path, my_annex_name)

    my_sender = '1659026978@qq.com'
    my_password = 'wmnwyialdvgffcjb'
    # 接收人邮箱列表
    # my_receiver = ['1812883703@qq.com', '365490010@qq.com', '853514522@qq.com', '34578008@qq.com']
    # my_receiver = ['1812883703@qq.com']
    with open('receiver.txt', 'r', encoding='UTF-8') as f:
        my_receiver = [r.strip() for r in f.readlines()]
    print('接收人：', my_receiver)
    # 发送邮件
    send_email(my_sender, my_password, my_receiver, my_msg)


if __name__ == "__main__":
    main()
